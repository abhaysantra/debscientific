from django.shortcuts import render, redirect
from django.http import JsonResponse, HttpResponse
from django.conf import settings  
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.utils.decorators import method_decorator
from datetime import datetime
import json, os,sys
from django.template import Context
import openpyxl
from . import accounts, deb_utility
from debadmin.models import pincode_service

def pincode_service_listview(request):
    if 'admin_session_id' not in request.session:
        return redirect(accounts.login)
    pincode_service_objs = pincode_service.objects.all()
    context = {'pincode_service_objs':pincode_service_objs, 'pincode_service_active':'active'}
    return render(request, 'pincode_service/pincode_service_view.html',context)

def pincode_service_add_view(request):
    context = {'pincode_service_active':'active'}
    return render(request, 'pincode_service/pincode_service_add_view.html',context)

def pincode_service_edit_view(request, pincode_service_id):
    pincode_service_det=pincode_service.objects.filter(id=pincode_service_id)
    context = {'pincode_service_det':pincode_service_det,'pincode_service_active':'active'}
    return render(request, 'pincode_service/pincode_service_edit_view.html',context)

def pincode_service_insert(request):

    pincode = request.POST["pincode"]
    # pincode_service_info = pincode_service(        
    #             pincode=pincode,
    #             status='active',
    #             )
    # pincode_service_info.save()

    obj, created = pincode_service.objects.get_or_create(pincode=pincode, status='active')
    messages.success(request,'Pincode Already Present in the List!')
    if created:
        messages.success(request,'Pincode Added Successfully!')
    return redirect(pincode_service_listview)

def pincode_service_status_change(request):
    status = request.POST['status']
    pincode_service_id = request.POST.getlist('id[]')
    
    for pincode_service_id in pincode_service_id:
        pincode_service_data = pincode_service.objects.filter(id=pincode_service_id)[0]
        pincode_service_data.status = status
        pincode_service_data.save(update_fields=['status'])


    return JsonResponse({'result': '1'})

def pincode_availability_check(request):
    pincode = request.POST["pincode"]
    
    pincode_obj = pincode_service.objects.filter(pincode=pincode)
    status = 'Pincode Not Exists'
    if pincode_obj:
        status = 'Pincode Already Exists'
    return JsonResponse({'result': status})

def pincode_service_delete(request, pincode_service_id):
    print("pincode_service_id id: ",pincode_service_id)
    pincode_service_det=pincode_service.objects.filter(id=pincode_service_id)[0].delete()

    messages.success(request,'Pincode Has Been Deleted Successfully!')
    return redirect(pincode_service_listview)

def pincode_service_add_bulk_pincode_view(request):
    context = {'pincode_service_active':'active'}
    return render(request, 'pincode_service/pincode_service_add_bulk_pincode_view.html',context)

def pincode_service_add_bulk_pincode(request):
    pincode_file = request.FILES["pincode_file"]

    try:
        wb_obj = openpyxl.load_workbook(pincode_file) 
        sheet_obj = wb_obj.active 
        m_row = sheet_obj.max_row 

        # all values of first column except 1st row as it contains pincode word: 
        for i in range(2, m_row + 1): 
            cell_obj = sheet_obj.cell(row = i, column = 1) 
            if cell_obj.value == '':
                break
            # print(cell_obj.value)
            pincode = int(str(cell_obj.value).strip())
            # pincode_service_info = pincode_service(        
            #             pincode=pincode,
            #             status='active',
            #             )
            # pincode_service_info.save()
            obj, created = pincode_service.objects.get_or_create(pincode=pincode, status='active')

        messages.success(request,'Pincode Added Successfully!')
    except Exception as e:
        messages.error(request,'External Pincode file is not formatted properly')
    return redirect(pincode_service_listview)


